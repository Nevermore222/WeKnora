import contextlib
import io
import json
import os
import subprocess
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl

import officecli_bridge


class OfficeCLIHelpersTest(unittest.TestCase):
    def test_resolve_relative_path_rejects_absolute_and_parent_escape(self):
        self.assertEqual(
            officecli_bridge.resolve_relative_path("report.docx"),
            "report.docx",
        )
        with self.assertRaises(ValueError):
            officecli_bridge.resolve_relative_path("../escape.docx")
        with self.assertRaises(ValueError):
            officecli_bridge.resolve_relative_path("/tmp/escape.docx")

    def test_props_to_cli_args_expands_sorted_prop_flags(self):
        args = officecli_bridge.props_to_cli_args({"text": "Hello", "x": "1cm"})
        self.assertEqual(args, ["--prop", "text=Hello", "--prop", "x=1cm"])

    def test_build_view_command_includes_mode_and_optional_flags(self):
        cmd, temp_path = officecli_bridge.build_officecli_command(
            {
                "action": "view",
                "file": "deck.pptx",
                "mode": "text",
                "page": "1-2",
                "max_lines": 20,
            },
            "request.json",
        )
        self.assertIsNone(temp_path)
        self.assertEqual(
            cmd,
            [
                "officecli",
                "view",
                "deck.pptx",
                "text",
                "--page",
                "1-2",
                "--max-lines",
                "20",
                "--json",
            ],
        )

    def test_build_officecli_passthrough_preserves_supported_options(self):
        cmd = officecli_bridge.build_officecli_passthrough_command(
            {
                "action": "officecli",
                "file": "brief.docx",
                "command": ["view", "{file}", "text", "--start", "2", "--end", "8"],
            }
        )
        self.assertEqual(
            cmd,
            [
                "officecli",
                "view",
                "brief.docx",
                "text",
                "--start",
                "2",
                "--end",
                "8",
                "--json",
            ],
        )

    def test_build_officecli_passthrough_resolves_declared_path_placeholders(self):
        cmd = officecli_bridge.build_officecli_passthrough_command(
            {
                "action": "officecli",
                "file": "brief.docx",
                "command": ["view", "{file}", "html", "--out", "{preview}"],
                "paths": {"preview": "previews/brief.html"},
            }
        )
        self.assertEqual(
            cmd,
            [
                "officecli",
                "view",
                "brief.docx",
                "html",
                "--out",
                "previews/brief.html",
                "--json",
            ],
        )

    def test_build_officecli_passthrough_rejects_non_document_management_verbs(self):
        with self.assertRaises(ValueError):
            officecli_bridge.build_officecli_passthrough_command(
                {
                    "action": "officecli",
                    "command": ["mcp", "register"],
                }
            )
        with self.assertRaises(ValueError):
            officecli_bridge.build_officecli_passthrough_command(
                {
                    "action": "officecli",
                    "file": "brief.docx",
                    "command": ["view", "../escape.docx", "text"],
                }
            )

    def test_run_officecli_uses_utf8_locale(self):
        with mock.patch.object(officecli_bridge.subprocess, "run") as mocked_run:
            mocked_run.return_value = subprocess.CompletedProcess([], 0, stdout="", stderr="")

            officecli_bridge.run_officecli(["officecli", "--version"])

        env = mocked_run.call_args.kwargs["env"]
        self.assertEqual(env["LANG"], "C.UTF-8")
        self.assertEqual(env["LC_ALL"], "C.UTF-8")

    def test_commit_staged_file_falls_back_to_copy_on_replace_permission_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original = root / "brief.xlsx"
            staged = root / ".brief.xelora-123.xlsx"
            original.write_bytes(b"original")
            staged.write_bytes(b"updated")

            with mock.patch.object(officecli_bridge.os, "replace", side_effect=PermissionError("denied")):
                method = officecli_bridge.commit_staged_file(staged, original, root)

            self.assertEqual(method, "copy")
            self.assertEqual(original.read_bytes(), b"updated")

    def test_commit_staged_file_reports_office_lock_and_keeps_pending_copy(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original = root / "brief.xlsx"
            staged = root / ".brief.xelora-123.xlsx"
            lock = root / "~$brief.xlsx"
            original.write_bytes(b"original")
            staged.write_bytes(b"updated")
            lock.write_bytes(b"")

            with mock.patch.object(officecli_bridge.os, "replace", side_effect=PermissionError("denied")):
                with self.assertRaises(PermissionError) as raised:
                    officecli_bridge.commit_staged_file(staged, original, root)

            self.assertIn("Close the workbook in Office/WPS", str(raised.exception))
            self.assertIn("~$brief.xlsx", str(raised.exception))
            pending = root / "brief.xelora-pending.xlsx"
            self.assertTrue(pending.exists())
            self.assertEqual(pending.read_bytes(), b"updated")
            self.assertEqual(original.read_bytes(), b"original")

    def test_mutation_validates_before_replacing_original(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original = root / "brief.docx"
            original.write_bytes(b"original")
            request = root / ".xelora" / "jobs" / "job-1" / "request.json"
            request.parent.mkdir(parents=True)
            request.write_text(
                json.dumps(
                    {
                        "action": "set",
                        "file": "brief.docx",
                        "path": "/body/p[1]",
                        "props": {"text": "Updated"},
                    }
                ),
                encoding="utf-8",
            )

            calls = []

            def fake_run(command, **kwargs):
                calls.append(command)
                if command[1] == "set":
                    (root / command[2]).write_bytes(b"updated")
                return subprocess.CompletedProcess(command, 0, stdout=f"updated {command[2]}", stderr="")

            previous_cwd = Path.cwd()
            try:
                os.chdir(root)
                with mock.patch.object(officecli_bridge.subprocess, "run", side_effect=fake_run):
                    output = io.StringIO()
                    with contextlib.redirect_stdout(output):
                        result = officecli_bridge.run_request(".xelora/jobs/job-1/request.json")
            finally:
                os.chdir(previous_cwd)

            self.assertEqual(result, 0)
            self.assertEqual(original.read_bytes(), b"updated")
            self.assertEqual(calls[0][1], "set")
            self.assertEqual(calls[1][1], "validate")
            self.assertNotEqual(calls[0][2], "brief.docx")
            self.assertIn("brief.docx", output.getvalue())
            self.assertNotIn(".xelora-", output.getvalue())

    def test_failed_validation_preserves_original(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original = root / "brief.docx"
            original.write_bytes(b"original")
            request = root / "request.json"
            request.write_text(
                json.dumps(
                    {
                        "action": "set",
                        "file": "brief.docx",
                        "path": "/body/p[1]",
                        "props": {"text": "Broken"},
                    }
                ),
                encoding="utf-8",
            )

            def fake_run(command, **kwargs):
                if command[1] == "set":
                    (root / command[2]).write_bytes(b"broken")
                    return subprocess.CompletedProcess(command, 0, stdout="{}", stderr="")
                return subprocess.CompletedProcess(command, 1, stdout="", stderr="invalid package")

            previous_cwd = Path.cwd()
            try:
                os.chdir(root)
                with mock.patch.object(officecli_bridge.subprocess, "run", side_effect=fake_run):
                    result = officecli_bridge.run_request("request.json")
            finally:
                os.chdir(previous_cwd)

            self.assertNotEqual(result, 0)
            self.assertEqual(original.read_bytes(), b"original")

    def test_officecli_passthrough_mutation_stages_and_validates(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            original = root / "brief.docx"
            original.write_bytes(b"original")
            request = root / "request.json"
            request.write_text(
                json.dumps(
                    {
                        "action": "officecli",
                        "file": "brief.docx",
                        "command": ["raw-set", "{file}", "/document", "--xml", "<w:document/>"],
                    }
                ),
                encoding="utf-8",
            )

            calls = []

            def fake_run(command, **kwargs):
                calls.append(command)
                if command[1] == "raw-set":
                    (root / command[2]).write_bytes(b"updated")
                return subprocess.CompletedProcess(command, 0, stdout=f"ok {command[2]}", stderr="")

            previous_cwd = Path.cwd()
            try:
                os.chdir(root)
                with mock.patch.object(officecli_bridge.subprocess, "run", side_effect=fake_run):
                    output = io.StringIO()
                    with contextlib.redirect_stdout(output):
                        result = officecli_bridge.run_request("request.json")
            finally:
                os.chdir(previous_cwd)

            self.assertEqual(result, 0)
            self.assertEqual(original.read_bytes(), b"updated")
            self.assertEqual([call[1] for call in calls], ["raw-set", "validate"])
            self.assertNotEqual(calls[0][2], "brief.docx")
            self.assertIn("brief.docx", output.getvalue())
            self.assertNotIn(".xelora-", output.getvalue())

    def test_write_docx_uses_compact_paragraph_request(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request = root / ".xelora" / "jobs" / "job-1" / "request.json"
            request.parent.mkdir(parents=True)
            request.write_text(
                json.dumps(
                    {
                        "action": "write_docx",
                        "file": "classic.docx",
                        "title": "三字经",
                        "paragraphs": ["人之初，性本善。", "戒之哉，宜勉力。"],
                        "force": True,
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            calls = []

            def fake_run(command, **kwargs):
                calls.append(command)
                if command[1] in {"create", "add"}:
                    (root / command[2]).write_bytes(b"docx")
                return subprocess.CompletedProcess(command, 0, stdout="{}", stderr="")

            previous_cwd = Path.cwd()
            try:
                os.chdir(root)
                with mock.patch.object(officecli_bridge.subprocess, "run", side_effect=fake_run):
                    result = officecli_bridge.run_request(".xelora/jobs/job-1/request.json")
            finally:
                os.chdir(previous_cwd)

            self.assertEqual(result, 0)
            self.assertEqual([call[1] for call in calls], ["create", "add", "add", "add", "validate"])
            self.assertEqual(calls[0][2], calls[1][2])
            self.assertNotEqual(calls[0][2], "classic.docx")
            self.assertEqual((root / "classic.docx").read_bytes(), b"docx")

    def test_write_xlsx_creates_workbook_from_sheets_request(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            request = root / ".xelora" / "jobs" / "job-1" / "request.json"
            request.parent.mkdir(parents=True)
            request.write_text(
                json.dumps(
                    {
                        "action": "write_xlsx",
                        "file": "summary.xlsx",
                        "sheets": [
                            {
                                "name": "\u9a8c\u8bc1",
                                "headers": ["\u9879\u76ee", "\u72b6\u6001"],
                                "rows": [["Chrome DevTools MCP", "\u5df2\u4fee\u590d"]],
                            }
                        ],
                        "force": True,
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            previous_cwd = Path.cwd()
            try:
                os.chdir(root)
                result = officecli_bridge.run_request(".xelora/jobs/job-1/request.json")
            finally:
                os.chdir(previous_cwd)

            workbook_path = root / "summary.xlsx"
            self.assertEqual(result, 0)
            self.assertTrue(workbook_path.exists())
            self.assertFalse(list(root.glob(".summary.xelora-*.xlsx")))

            workbook = openpyxl.load_workbook(workbook_path)
            try:
                worksheet = workbook["\u9a8c\u8bc1"]
                self.assertEqual(worksheet["A1"].value, "\u9879\u76ee")
                self.assertEqual(worksheet["B1"].value, "\u72b6\u6001")
                self.assertEqual(worksheet["A2"].value, "Chrome DevTools MCP")
                self.assertEqual(worksheet["B2"].value, "\u5df2\u4fee\u590d")
                self.assertEqual(worksheet.freeze_panes, "A2")
                self.assertTrue(worksheet["A1"].font.bold)
            finally:
                workbook.close()

    def test_run_python_can_apply_xlsx_styles_atomically(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "styled.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(["Section", "Title"])
            worksheet.append(["Dao", "Chapter 1"])
            workbook.save(workbook_path)
            workbook.close()

            request = root / "request.json"
            request.write_text(
                json.dumps(
                    {
                        "action": "run_python",
                        "file": "styled.xlsx",
                        "code": "\n".join(
                            [
                                "from openpyxl import load_workbook",
                                "from openpyxl.styles import Font, PatternFill, Border, Side",
                                "wb = load_workbook(target_file)",
                                "ws = wb.active",
                                "ws['A1'].fill = PatternFill('solid', fgColor='4A0E4E')",
                                "ws['A1'].font = Font(color='FFFFFF', bold=True)",
                                "thin = Side(style='thin', color='000000')",
                                "ws['A1'].border = Border(left=thin, right=thin, top=thin, bottom=thin)",
                                "wb.save(target_file)",
                            ]
                        ),
                    }
                ),
                encoding="utf-8",
            )

            previous_cwd = Path.cwd()
            try:
                os.chdir(root)
                result = officecli_bridge.run_request("request.json")
            finally:
                os.chdir(previous_cwd)

            self.assertEqual(result, 0)
            self.assertFalse(list(root.glob(".styled.xelora-*.xlsx")))
            workbook = openpyxl.load_workbook(workbook_path)
            try:
                worksheet = workbook.active
                self.assertEqual(worksheet["A1"].fill.fgColor.rgb, "004A0E4E")
                self.assertEqual(worksheet["A1"].font.color.rgb, "00FFFFFF")
                self.assertTrue(worksheet["A1"].font.bold)
                self.assertEqual(worksheet["A1"].border.left.style, "thin")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
