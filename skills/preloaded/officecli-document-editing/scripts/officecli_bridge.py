#!/usr/bin/env python3
import json
import os
import shutil
import subprocess
import sys
import tempfile
import uuid
from pathlib import Path


MUTATING_ACTIONS = {
    "create",
    "save",
    "remove",
    "set",
    "add",
    "batch",
    "write_docx",
    "write_xlsx",
    "run_python",
}

OFFICECLI_PASSTHROUGH_VERBS = {
    "open",
    "close",
    "watch",
    "unwatch",
    "view",
    "get",
    "query",
    "set",
    "add",
    "remove",
    "move",
    "swap",
    "refresh",
    "raw",
    "raw-set",
    "add-part",
    "validate",
    "save",
    "batch",
    "dump",
    "import",
    "create",
    "merge",
    "help",
}

OFFICECLI_MUTATING_VERBS = {
    "close",
    "save",
    "set",
    "add",
    "remove",
    "move",
    "swap",
    "refresh",
    "raw-set",
    "add-part",
    "batch",
    "import",
    "create",
    "merge",
}


def resolve_relative_path(candidate):
    if candidate is None:
        raise ValueError("path is required")
    normalized = os.path.normpath(str(candidate).strip())
    if normalized in ("", ".", os.sep):
        raise ValueError(f"invalid relative path: {candidate}")
    if os.path.isabs(normalized) or normalized.startswith(".."):
        raise ValueError(f"invalid relative path: {candidate}")
    return normalized.replace("\\", "/")


def props_to_cli_args(props):
    if not props:
        return []
    args = []
    for key in sorted(props.keys()):
        value = props[key]
        args.extend(["--prop", f"{key}={value}"])
    return args


def require_field(payload, name):
    value = payload.get(name)
    if value in (None, ""):
        raise ValueError(f"{name} is required")
    return value


def build_officecli_command(payload, request_path):
    action = require_field(payload, "action")
    if action == "create":
        file_path = resolve_relative_path(require_field(payload, "file"))
        cmd = ["officecli", "create", file_path, "--json"]
        if payload.get("force"):
            cmd.append("--force")
        if payload.get("type"):
            cmd.extend(["--type", str(payload["type"])])
        return cmd, None

    file_path = resolve_relative_path(require_field(payload, "file"))

    if action == "validate":
        return ["officecli", "validate", file_path, "--json"], None
    if action == "save":
        return ["officecli", "save", file_path, "--json"], None
    if action == "close":
        return ["officecli", "close", file_path, "--json"], None
    if action == "view":
        mode = require_field(payload, "mode")
        cmd = ["officecli", "view", file_path, str(mode)]
        if payload.get("page") is not None:
            cmd.extend(["--page", str(payload["page"])])
        if payload.get("range") is not None:
            cmd.extend(["--range", str(payload["range"])])
        if payload.get("max_lines") is not None:
            cmd.extend(["--max-lines", str(payload["max_lines"])])
        if payload.get("out") is not None:
            cmd.extend(["--out", resolve_relative_path(payload["out"])])
        cmd.append("--json")
        return cmd, None
    if action == "get":
        path = payload.get("path", "/")
        cmd = ["officecli", "get", file_path, str(path), "--json"]
        if payload.get("depth") is not None:
            cmd.extend(["--depth", str(payload["depth"])])
        return cmd, None
    if action == "query":
        selector = require_field(payload, "selector")
        cmd = ["officecli", "query", file_path, str(selector), "--json"]
        if payload.get("find"):
            cmd.extend(["--find", str(payload["find"])])
        return cmd, None
    if action == "remove":
        path = require_field(payload, "path")
        return ["officecli", "remove", file_path, str(path), "--json"], None
    if action == "set":
        path = require_field(payload, "path")
        cmd = ["officecli", "set", file_path, str(path), "--json"]
        cmd.extend(props_to_cli_args(payload.get("props")))
        if payload.get("find"):
            cmd.extend(["--find", str(payload["find"])])
        if payload.get("replace") is not None:
            cmd.extend(["--replace", str(payload["replace"])])
        return cmd, None
    if action == "add":
        parent = require_field(payload, "parent")
        cmd = ["officecli", "add", file_path, str(parent), "--json"]
        if payload.get("type"):
            cmd.extend(["--type", str(payload["type"])])
        if payload.get("from"):
            cmd.extend(["--from", str(payload["from"])])
        if payload.get("index") is not None:
            cmd.extend(["--index", str(payload["index"])])
        if payload.get("after"):
            cmd.extend(["--after", str(payload["after"])])
        if payload.get("before"):
            cmd.extend(["--before", str(payload["before"])])
        cmd.extend(props_to_cli_args(payload.get("props")))
        return cmd, None
    if action == "batch":
        commands = payload.get("commands")
        if not isinstance(commands, list) or not commands:
            raise ValueError("commands must be a non-empty array for batch")
        fd, temp_path = tempfile.mkstemp(
            prefix="officecli-batch-",
            suffix=".json",
            dir=os.path.dirname(os.path.abspath(request_path)),
        )
        os.close(fd)
        with open(temp_path, "w", encoding="utf-8") as handle:
            json.dump(commands, handle, ensure_ascii=False)
        cmd = ["officecli", "batch", file_path, "--input", temp_path, "--json"]
        if payload.get("stop_on_error"):
            cmd.append("--stop-on-error")
        return cmd, temp_path

    raise ValueError(f"unsupported action: {action}")


def build_officecli_passthrough_command(payload):
    raw_command = payload.get("command")
    if raw_command is None:
        raw_command = payload.get("args")
    if not isinstance(raw_command, list) or not raw_command:
        raise ValueError("command must be a non-empty array for officecli")

    command = [str(item) for item in raw_command]
    verb = command[0]
    if verb == "officecli":
        raise ValueError("command must not include the officecli binary")
    if verb not in OFFICECLI_PASSTHROUGH_VERBS:
        raise ValueError(f"unsupported officecli command: {verb}")
    if verb != "help":
        if payload.get("file") is None:
            raise ValueError("file is required for officecli document commands")
        if "{file}" not in command:
            raise ValueError("officecli document commands must use the {file} placeholder")

    placeholders = {}
    if payload.get("file") is not None:
        placeholders["{file}"] = resolve_relative_path(payload["file"])

    declared_paths = payload.get("paths", {})
    if declared_paths is not None:
        if not isinstance(declared_paths, dict):
            raise ValueError("paths must be an object")
        for name, path_value in declared_paths.items():
            key = "{" + str(name) + "}"
            if key == "{file}":
                raise ValueError("paths must not override {file}")
            placeholders[key] = resolve_relative_path(path_value)

    resolved_args = []
    for token in command:
        replacement = placeholders.get(token)
        if replacement is not None:
            resolved_args.append(replacement)
            continue
        if token in {".."} or token.startswith("../") or token.startswith("..\\"):
            raise ValueError(f"invalid officecli argument path: {token}")
        if len(token) >= 3 and token[1:3] in {":\\", ":/"}:
            raise ValueError(f"invalid officecli argument path: {token}")
        resolved_args.append(token)

    if payload.get("json", True) and "--json" not in resolved_args:
        resolved_args.append("--json")
    return ["officecli", *resolved_args]


def officecli_passthrough_mutates(payload):
    if str(payload.get("action")) != "officecli":
        return False
    if payload.get("mutates") is not None:
        return bool(payload.get("mutates"))
    raw_command = payload.get("command")
    if raw_command is None:
        raw_command = payload.get("args")
    if not isinstance(raw_command, list) or not raw_command:
        return False
    return str(raw_command[0]) in OFFICECLI_MUTATING_VERBS


def is_mutating_payload(payload):
    action = str(require_field(payload, "action"))
    if action == "officecli":
        return officecli_passthrough_mutates(payload)
    return action in MUTATING_ACTIONS


def resolve_workspace_path(base_dir, candidate):
    relative = resolve_relative_path(candidate)
    resolved = (Path(base_dir) / relative).resolve()
    try:
        resolved.relative_to(Path(base_dir).resolve())
    except ValueError as exc:
        raise ValueError(f"path escapes conversation workspace: {candidate}") from exc
    return resolved


def print_completed(completed, replacements=None):
    replacements = replacements or {}

    def display(value):
        for source, target in replacements.items():
            value = value.replace(source, target)
        return value

    if completed.stdout:
        print(display(completed.stdout), end="")
    if completed.stderr:
        print(display(completed.stderr), file=sys.stderr, end="")


def run_officecli(command):
    env = os.environ.copy()
    env["LANG"] = "C.UTF-8"
    env["LC_ALL"] = "C.UTF-8"
    return subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False,
        env=env,
    )


def office_lock_file_for(path):
    return path.with_name(f"~${path.name}")


def unique_sibling_path(path, label):
    candidate = path.with_name(f"{path.stem}.{label}{path.suffix}")
    if not candidate.exists():
        return candidate
    for index in range(2, 100):
        candidate = path.with_name(f"{path.stem}.{label}-{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise ValueError(f"could not allocate a unique {label} path for {path.name}")


def commit_staged_file(temp_path, final_path, base_dir):
    try:
        os.replace(temp_path, final_path)
        return "replace"
    except PermissionError as exc:
        lock_file = office_lock_file_for(final_path)
        if lock_file.exists():
            pending_path = unique_sibling_path(final_path, "xelora-pending")
            shutil.copy2(temp_path, pending_path)
            raise PermissionError(
                "target file is locked and could not be replaced. "
                "Close the workbook in Office/WPS and retry. "
                f"Lock file: {lock_file.relative_to(base_dir).as_posix()}. "
                f"Validated pending output: {pending_path.relative_to(base_dir).as_posix()}."
            ) from exc

        # Some Windows bind mounts reject rename-over-existing while allowing a
        # direct overwrite. The staged file has already been validated.
        shutil.copy2(temp_path, final_path)
        return "copy"


def run_read_only(payload, request_path):
    if str(payload.get("action")) == "officecli":
        completed = run_officecli(build_officecli_passthrough_command(payload))
        print_completed(completed)
        return completed.returncode

    command, temp_path = build_officecli_command(payload, str(request_path))
    try:
        completed = run_officecli(command)
        print_completed(completed)
        return completed.returncode
    finally:
        if temp_path:
            Path(temp_path).unlink(missing_ok=True)


def run_mutation(payload, request_path, base_dir):
    final_path = resolve_workspace_path(base_dir, require_field(payload, "file"))
    final_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = final_path.with_name(
        f".{final_path.stem}.xelora-{uuid.uuid4().hex}{final_path.suffix}"
    )
    if final_path.exists():
        shutil.copy2(final_path, temp_path)

    staged_payload = dict(payload)
    staged_payload["file"] = temp_path.relative_to(base_dir).as_posix()
    display_file = final_path.relative_to(base_dir).as_posix()
    replacements = {staged_payload["file"]: display_file}
    if str(payload.get("action")) == "write_docx":
        return run_write_docx_mutation(staged_payload, temp_path, final_path, base_dir, replacements)
    if str(payload.get("action")) == "write_xlsx":
        return run_write_xlsx_mutation(staged_payload, temp_path, final_path, base_dir, replacements)
    if str(payload.get("action")) == "run_python":
        return run_python_mutation(staged_payload, temp_path, final_path, base_dir, replacements, str(request_path))
    if str(payload.get("action")) == "officecli":
        command = build_officecli_passthrough_command(staged_payload)
        try:
            completed = run_officecli(command)
            print_completed(completed, replacements)
            if completed.returncode != 0:
                return completed.returncode

            validated = run_officecli(
                ["officecli", "validate", staged_payload["file"], "--json"]
            )
            print_completed(validated, replacements)
            if validated.returncode != 0:
                print("office_validation_failed", file=sys.stderr)
                return validated.returncode or 1

            commit_staged_file(temp_path, final_path, base_dir)
            return 0
        finally:
            temp_path.unlink(missing_ok=True)

    command, batch_path = build_officecli_command(staged_payload, str(request_path))
    try:
        completed = run_officecli(command)
        print_completed(completed, replacements)
        if completed.returncode != 0:
            return completed.returncode

        validated = run_officecli(
            ["officecli", "validate", staged_payload["file"], "--json"]
        )
        print_completed(validated, replacements)
        if validated.returncode != 0:
            print("office_validation_failed", file=sys.stderr)
            return validated.returncode or 1

        commit_staged_file(temp_path, final_path, base_dir)
        return 0
    finally:
        if batch_path:
            Path(batch_path).unlink(missing_ok=True)
        temp_path.unlink(missing_ok=True)


def normalize_paragraphs(payload):
    raw = payload.get("paragraphs")
    if raw is None:
        content = payload.get("content")
        if isinstance(content, str):
            raw = [line.strip() for line in content.splitlines() if line.strip()]
    if not isinstance(raw, list) or not raw:
        raise ValueError("paragraphs must be a non-empty array for write_docx")
    paragraphs = []
    for item in raw:
        text = str(item).strip()
        if text:
            paragraphs.append(text)
    if not paragraphs:
        raise ValueError("paragraphs must contain at least one non-empty paragraph")
    return paragraphs


def run_write_docx_mutation(payload, temp_path, final_path, base_dir, replacements):
    if final_path.suffix.lower() != ".docx":
        raise ValueError("write_docx requires a .docx file")

    title = str(payload.get("title", "")).strip()
    paragraphs = normalize_paragraphs(payload)
    staged_file = temp_path.relative_to(base_dir).as_posix()

    commands = [["officecli", "create", staged_file, "--json", "--force"]]
    if title:
        commands.append(
            [
                "officecli",
                "add",
                staged_file,
                "/",
                "--type",
                "paragraph",
                "--prop",
                f"text={title}",
                "--prop",
                "style=Title",
                "--json",
            ]
        )
    for paragraph in paragraphs:
        commands.append(
            [
                "officecli",
                "add",
                staged_file,
                "/",
                "--type",
                "paragraph",
                "--prop",
                f"text={paragraph}",
                "--json",
            ]
        )
    commands.append(["officecli", "validate", staged_file, "--json"])

    try:
        for command in commands:
            completed = run_officecli(command)
            print_completed(completed, replacements)
            if completed.returncode != 0:
                return completed.returncode
        commit_staged_file(temp_path, final_path, base_dir)
        return 0
    finally:
        temp_path.unlink(missing_ok=True)


def normalize_xlsx_sheets(payload):
    raw_sheets = payload.get("sheets")
    if raw_sheets is None:
        raw_sheets = [
            {
                "name": payload.get("sheet", "Sheet1"),
                "headers": payload.get("headers", []),
                "rows": payload.get("rows", []),
            }
        ]
    if not isinstance(raw_sheets, list) or not raw_sheets:
        raise ValueError("sheets must be a non-empty array for write_xlsx")

    sheets = []
    for index, raw_sheet in enumerate(raw_sheets, start=1):
        if not isinstance(raw_sheet, dict):
            raise ValueError("each sheet must be an object")
        name = str(raw_sheet.get("name") or f"Sheet{index}").strip()
        if not name:
            name = f"Sheet{index}"
        if len(name) > 31 or any(char in name for char in "[]:*?/\\"):
            raise ValueError(f"invalid sheet name: {name}")

        headers = raw_sheet.get("headers", [])
        rows = raw_sheet.get("rows", [])
        if not isinstance(headers, list):
            raise ValueError("headers must be an array")
        if not isinstance(rows, list):
            raise ValueError("rows must be an array")
        normalized_rows = []
        for row in rows:
            if not isinstance(row, list):
                raise ValueError("each row must be an array")
            normalized_rows.append(row)
        if not headers and not normalized_rows:
            raise ValueError("each sheet must include headers or rows")
        sheets.append({"name": name, "headers": headers, "rows": normalized_rows})
    return sheets


def xlsx_cell_value(value):
    if value is None or isinstance(value, (bool, int, float)):
        return value
    return str(value)


def run_write_xlsx_mutation(payload, temp_path, final_path, base_dir, replacements):
    if final_path.suffix.lower() != ".xlsx":
        raise ValueError("write_xlsx requires a .xlsx file")

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    sheets = normalize_xlsx_sheets(payload)
    workbook = Workbook()
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for index, sheet in enumerate(sheets):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = sheet["name"]
        headers = [xlsx_cell_value(value) for value in sheet["headers"]]
        if headers:
            worksheet.append(headers)
        for row in sheet["rows"]:
            worksheet.append([xlsx_cell_value(value) for value in row])

        if headers:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
        for row in worksheet.iter_rows(min_row=2 if headers else 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)

    try:
        workbook.save(temp_path)
        workbook.close()
        validated = load_workbook(temp_path, read_only=True)
        validated.close()
        commit_staged_file(temp_path, final_path, base_dir)
        display_file = final_path.relative_to(base_dir).as_posix()
        print(
            json.dumps(
                {
                    "status": "success",
                    "file": display_file,
                    "sheets": [sheet["name"] for sheet in sheets],
                },
                ensure_ascii=False,
            )
        )
        return 0
    finally:
        temp_path.unlink(missing_ok=True)


def validate_python_office_file(path):
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True)
        workbook.close()
        return
    if suffix in {".docx", ".pptx"}:
        import zipfile

        if not zipfile.is_zipfile(path):
            raise ValueError(f"{suffix} output is not a valid Office package")


def run_python_mutation(payload, temp_path, final_path, base_dir, replacements, request_path):
    code = payload.get("code")
    if not isinstance(code, str) or not code.strip():
        raise ValueError("code must be a non-empty string for run_python")

    script_path = Path(request_path).with_name(f"office-python-{uuid.uuid4().hex}.py")
    wrapper = "\n".join(
        [
            "import json",
            "from pathlib import Path",
            f"workspace_dir = Path({str(base_dir)!r})",
            f"target_file = Path({str(temp_path)!r})",
            f"final_file = Path({str(final_path)!r})",
            f"payload = json.loads({json.dumps(json.dumps(payload, ensure_ascii=False))})",
            code,
            "",
        ]
    )
    try:
        script_path.write_text(wrapper, encoding="utf-8")
        completed = run_officecli([sys.executable, script_path.relative_to(base_dir).as_posix()])
        print_completed(completed, replacements)
        if completed.returncode != 0:
            return completed.returncode
        if not temp_path.exists():
            raise ValueError("run_python did not create or update the target file")
        validate_python_office_file(temp_path)
        commit_staged_file(temp_path, final_path, base_dir)
        print(
            json.dumps(
                {
                    "status": "success",
                    "file": final_path.relative_to(base_dir).as_posix(),
                    "action": "run_python",
                },
                ensure_ascii=False,
            )
        )
        return 0
    finally:
        script_path.unlink(missing_ok=True)
        temp_path.unlink(missing_ok=True)


def run_request(request_arg):
    base_dir = Path.cwd().resolve()
    request_path = resolve_workspace_path(base_dir, request_arg)

    with request_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("request payload must be a JSON object")

    if is_mutating_payload(payload):
        return run_mutation(payload, request_path, base_dir)
    return run_read_only(payload, request_path)


def main():
    if len(sys.argv) < 2:
        print("usage: officecli_bridge.py <request.json>", file=sys.stderr)
        return 2
    return run_request(sys.argv[1])


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1)
