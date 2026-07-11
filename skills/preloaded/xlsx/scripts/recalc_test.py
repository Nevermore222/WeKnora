import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

import recalc


class RecalcTest(unittest.TestCase):
    def test_no_formula_workbook_does_not_require_libreoffice(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "plain.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Name"
            ws["B1"] = "Value"
            ws["A2"] = "Chrome DevTools MCP"
            ws["B2"] = "Knowledge summary"
            wb.save(path)
            wb.close()

            with mock.patch.object(recalc, "setup_libreoffice_macro") as setup_macro:
                result = recalc.recalc(str(path))

            setup_macro.assert_not_called()
            self.assertEqual(result["status"], "success")
            self.assertEqual(result["total_errors"], 0)
            self.assertEqual(result["total_formulas"], 0)


if __name__ == "__main__":
    unittest.main()
